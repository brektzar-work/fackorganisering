"""
Excel-exportmodul för Vision Sektion 10

Denna modul skapar en Excel-fil med all information från systemet.
Filen innehåller flera flikar:
- Översikt: En hierarkisk vy över hela organisationen
- Personer: Lista på alla personer med deras roller och kontaktuppgifter
- Arbetsplatser: Information om alla arbetsplatser och deras placering
- Förvaltningar: Översikt över förvaltningar och deras statistik
- Avdelningar: Lista över avdelningar och deras information
- Enheter: Information om enheter och deras medlemmar
- Styrelser & Nämnder: Lista över alla styrelser med representanter

Tekniska detaljer:
- Använder openpyxl för att skapa och formatera Excel-filer
- Sparar filen i minnet för direkt nedladdning
- Formaterar data med Vision's färgschema
- Anpassar kolumnbredder automatiskt för bättre läsbarhet
"""

import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side
from views.cache_manager import get_cached_data, update_cache_after_change


def create_overview_sheet(writer, cached_data):
    """
    Skapar översiktsbladet med organisationsstruktur och färgkodning.
    
    Detta blad visar:
    - Förvaltningar i grönt
    - Avdelningar i mörklila
    - Enheter i lila
    - Arbetsplatser i ljuslila
    - Visionombud och skyddsombud med egna färger
    """
    # Skapa arbetsbok och lägg till blad
    workbook = writer.book
    sheet = workbook.create_sheet("Översikt")
    
    # Definiera stilar
    styles = {
        'forvaltning': {
            'fill': openpyxl.styles.PatternFill(start_color='00A68A', end_color='00A68A', fill_type='solid'),
            'font': openpyxl.styles.Font(color='EFE9E5', bold=True, size=12),
            'border': openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style='thin', color='210061'),
                top=openpyxl.styles.Side(style='medium', color='210061')
            ),
        },
        'avdelning': {
            'fill': openpyxl.styles.PatternFill(start_color='210061', end_color='210061', fill_type='solid'),
            'font': openpyxl.styles.Font(color='EFE9E5', bold=True, size=11),
            'border': openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style='thin', color='210061'),
                top=openpyxl.styles.Side(style='medium', color='210061')
            ),
        },
        'enhet': {
            'fill': openpyxl.styles.PatternFill(start_color='7930AE', end_color='7930AE', fill_type='solid'),
            'font': openpyxl.styles.Font(color='EFE9E5', bold=True),
            'border': openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style='thin', color='210061'),
                top=openpyxl.styles.Side(style='medium', color='210061')
            ),
        },
        'arbetsplats': {
            'fill': openpyxl.styles.PatternFill(start_color='a16ec6', end_color='a16ec6', fill_type='solid'),
            'font': openpyxl.styles.Font(color='EFE9E5'),
            'border': openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style='thin', color='7930AE'),
                top=openpyxl.styles.Side(style='medium', color='7930AE')
            ),
        },
        'visionombud': {
            'fill': openpyxl.styles.PatternFill(start_color='bc98d7', end_color='bc98d7', fill_type='solid'),
            'font': openpyxl.styles.Font(color='EFE9E5', italic=True),
            'border': openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style='thin', color='a16ec6'),
                top=openpyxl.styles.Side(style='medium', color='a16ec6')
            ),
        },
        'skyddsombud': {
            'fill': openpyxl.styles.PatternFill(start_color='c9acdf', end_color='c9acdf', fill_type='solid'),
            'font': openpyxl.styles.Font(color='EFE9E5', italic=True),
            'border': openpyxl.styles.Border(
                bottom=openpyxl.styles.Side(style='thin', color='a16ec6'),
                top=openpyxl.styles.Side(style='medium', color='a16ec6')
            ),
        }
    }
    
    # Add headers
    headers = ['Förvaltning', 'Avdelning', 'Enhet', 'Arbetsplats', 'Ombud']
    header_style = {
        'fill': openpyxl.styles.PatternFill(start_color='00A68A', end_color='00A68A', fill_type='solid'),
        'font': openpyxl.styles.Font(color='EFE9E5', bold=True, size=12)
    }
    
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = header_style['font']
        cell.fill = header_style['fill']
        cell.border = openpyxl.styles.Border(
            bottom=openpyxl.styles.Side(style='medium', color='EFE9E5')
        )
    
    # Start writing data
    row = 2
    
    # Sort förvaltningar by name
    forvaltningar = sorted(cached_data['forvaltningar'], key=lambda x: x['namn'])
    
    for forv in forvaltningar:
        # Write förvaltning
        for col in range(1, 6):
            cell = sheet.cell(row=row, column=col)
            cell.fill = styles['forvaltning']['fill']
            cell.font = styles['forvaltning']['font']
            cell.border = styles['forvaltning']['border']
            if col == 1:
                cell.value = forv['namn']
        row += 1
        
        # Get and sort avdelningar for this förvaltning
        avdelningar = sorted(
            [a for a in cached_data['avdelningar'] if str(a['forvaltning_id']) == str(forv['_id'])], 
            key=lambda x: x['namn']
        )
        
        for avd in avdelningar:
            # Write avdelning
            for col in range(1, 6):
                cell = sheet.cell(row=row, column=col)
                cell.fill = styles['avdelning']['fill']
                cell.font = styles['avdelning']['font']
                cell.border = styles['avdelning']['border']
                if col == 2:
                    cell.value = avd['namn']
            row += 1
            
            # Get and sort enheter for this avdelning
            enheter = sorted(
                [e for e in cached_data['enheter'] if str(e['avdelning_id']) == str(avd['_id'])], 
                key=lambda x: x['namn']
            )
            
            for enhet in enheter:
                # Write enhet
                for col in range(1, 6):
                    cell = sheet.cell(row=row, column=col)
                    cell.fill = styles['enhet']['fill']
                    cell.font = styles['enhet']['font']
                    cell.border = styles['enhet']['border']
                    if col == 3:
                        cell.value = enhet['namn']
                row += 1
                
                # Get arbetsplatser from the enhet's arbetsplatser array
                arbetsplatser = []
                if enhet.get('arbetsplatser'):
                    arbetsplatser = sorted(
                        [a for a in cached_data['arbetsplatser'] 
                         if str(a['_id']) in enhet['arbetsplatser']], 
                        key=lambda x: x['namn']
                    )
                
                # If no arbetsplatser, show placeholder
                if not arbetsplatser:
                    arbetsplatser = [{'namn': 'Ingen arbetsplats'}]
                
                for arbetsplats in arbetsplatser:
                    # Write arbetsplats
                    for col in range(1, 6):
                        cell = sheet.cell(row=row, column=col)
                        cell.fill = styles['arbetsplats']['fill']
                        cell.font = styles['arbetsplats']['font']
                        cell.border = styles['arbetsplats']['border']
                        if col == 4:
                            cell.value = arbetsplats['namn']
                    row += 1
                    
                    # Get and sort personer for this arbetsplats
                    personer = []
                    if arbetsplats['namn'] != 'Ingen arbetsplats':
                        personer = sorted(
                            [p for p in cached_data['personer'] 
                             if arbetsplats['namn'] in p.get('arbetsplats', [])],
                            key=lambda x: x['namn']
                        )
                    
                    # Write Visionombud
                    visionombud = [p for p in personer if p.get('visionombud')]
                    for person in visionombud:
                        for col in range(1, 6):
                            cell = sheet.cell(row=row, column=col)
                            cell.fill = styles['visionombud']['fill']
                            cell.font = styles['visionombud']['font']
                            cell.border = styles['visionombud']['border']
                            if col == 5:
                                cell.value = f"Visionombud: {person['namn']}"
                        row += 1
                    
                    # Write Skyddsombud
                    skyddsombud = [p for p in personer if p.get('skyddsombud')]
                    for person in skyddsombud:
                        for col in range(1, 6):
                            cell = sheet.cell(row=row, column=col)
                            cell.fill = styles['skyddsombud']['fill']
                            cell.font = styles['skyddsombud']['font']
                            cell.border = styles['skyddsombud']['border']
                            if col == 5:
                                cell.value = f"Skyddsombud: {person['namn']}"
                        row += 1
    
    # Adjust column widths
    for col in range(1, 6):
        max_length = 0
        for cell in sheet[openpyxl.utils.get_column_letter(col)]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = max_length + 2


def get_fresh_data(db):
    """
    Hämtar aktuell data direkt från databasen.
    Används för att säkerställa att exporten innehåller senaste informationen.
    """
    return {
        'forvaltningar': list(db.forvaltningar.find()),
        'avdelningar': list(db.avdelningar.find()),
        'enheter': list(db.enheter.find()),
        'arbetsplatser': list(db.arbetsplatser.find()),
        'personer': list(db.personer.find()),
        'boards': list(db.boards.find())
    }

def apply_sheet_styling(sheet, df):
    """
    Ger alla blad en enhetlig formatering.
    
    - Gröna rubriker med vit text
    - Varannan rad ljusgrå för bättre läsbarhet
    - Anpassar kolumnbredder efter innehåll (max 50 tecken)
    - Lägger till tunna linjer mellan raderna
    """
    # Stil för rubriker
    header_style = {
        'fill': PatternFill(start_color='00A68A', end_color='00A68A', fill_type='solid'),
        'font': Font(color='EFE9E5', bold=True, size=11),
        'border': Border(bottom=Side(style='medium', color='210061'))
    }
    
    # Stil för datarader
    row_style = {
        'even': {
            'fill': PatternFill(start_color='EFE9E5', end_color='EFE9E5', fill_type='solid'),
            'font': Font(color='210061'),
            'border': Border(
                bottom=Side(style='thin', color='210061'),
                top=Side(style='thin', color='210061')
            )
        },
        'odd': {
            'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
            'font': Font(color='210061'),
            'border': Border(
                bottom=Side(style='thin', color='210061'),
                top=Side(style='thin', color='210061')
            )
        }
    }
    
    # Applicera rubrikstil
    for col, header in enumerate(df.columns, 1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.fill = header_style['fill']
        cell.font = header_style['font']
        cell.border = header_style['border']
    
    # Applicera radstil och skriv data
    for row_idx, row in enumerate(df.values, 2):
        style = row_style['even'] if row_idx % 2 == 0 else row_style['odd']
        for col_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = style['fill']
            cell.font = style['font']
            cell.border = style['border']
    
    # Justera kolumnbredder
    for col in range(1, len(df.columns) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col)
        
        # Kontrollera rubrikbredd
        header_text = str(df.columns[col-1])
        max_length = len(header_text) + 2
        
        # Kontrollera databredd
        for cell in sheet[column_letter][1:]:
            try:
                cell_text = str(cell.value)
                max_length = max(max_length, len(cell_text) + 2)
            except:
                pass
        
        # Sätt bredd med max 50 tecken
        sheet.column_dimensions[column_letter].width = min(max_length, 50)


def create_excel_file(db):
    """
    Skapar en Excel-fil med all data från systemet.
    
    - Hämtar färsk data från databasen
    - Skapar alla blad med relevant information
    - Formaterar varje blad enligt Vision's grafiska profil
    - Returnerar filen som bytes för nedladdning
    """
    # Hämta färsk data
    data = get_fresh_data(db)
    
    # Skapa Excel-fil i minnet
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Skapa översiktsblad
        create_overview_sheet(writer, data)

        # Arbetsplatser
        df_arbetsplatser = pd.DataFrame([{
            'Namn': a['namn'],
            'Förvaltning': a.get('forvaltning_namn', ''),
            'Avdelning': a.get('avdelning_namn', ''),
            'Enhet': a.get('enhet_namn', ''),
            'Regional': 'Ja' if a.get('alla_forvaltningar', False) else 'Nej',
            'Adress': a.get('gatuadress', ''),
            'Postnummer': a.get('postnummer', ''),
            'Ort': a.get('ort', ''),
            'Kommun': a.get('kommun', ''),
            'Antal Medlemmar': len([p for p in data['personer'] 
                               if a['namn'] in p.get('arbetsplats', [])])
        } for a in data['arbetsplatser']])
        df_arbetsplatser.to_excel(writer, sheet_name='Arbetsplatser', index=False)
        apply_sheet_styling(writer.sheets['Arbetsplatser'], df_arbetsplatser)

        # Förvaltningar
        df_forvaltningar = pd.DataFrame([{
            'Namn': f['namn'],
            'Chef': f.get('chef', ''),
            'Antal Avdelningar': len([a for a in data['avdelningar'] 
                                 if str(a['forvaltning_id']) == str(f['_id'])]),
            'Antal Enheter': len([e for e in data['enheter'] 
                             if str(e['forvaltning_id']) == str(f['_id'])]),
            'Antal Arbetsplatser': len([a for a in data['arbetsplatser'] 
                                   if str(a.get('forvaltning_id')) == str(f['_id'])]),
            'Antal Medlemmar': len([p for p in data['personer'] 
                               if str(p['forvaltning_id']) == str(f['_id'])])
        } for f in data['forvaltningar']])
        df_forvaltningar.to_excel(writer, sheet_name='Förvaltningar', index=False)
        apply_sheet_styling(writer.sheets['Förvaltningar'], df_forvaltningar)

        # Avdelningar
        df_avdelningar = pd.DataFrame([{
            'Namn': a['namn'],
            'Förvaltning': a.get('forvaltning_namn', ''),
            'Chef': a.get('chef', ''),
            'Antal Enheter': len([e for e in data['enheter'] 
                             if str(e['avdelning_id']) == str(a['_id'])]),
            'Antal Medlemmar': len([p for p in data['personer'] 
                               if str(p['avdelning_id']) == str(a['_id'])])
        } for a in data['avdelningar']])
        df_avdelningar.to_excel(writer, sheet_name='Avdelningar', index=False)
        apply_sheet_styling(writer.sheets['Avdelningar'], df_avdelningar)

        # Enheter
        df_enheter = pd.DataFrame([{
            'Namn': e['namn'],
            'Förvaltning': e.get('forvaltning_namn', ''),
            'Avdelning': e.get('avdelning_namn', ''),
            'Chef': e.get('chef', ''),
            'Antal Medlemmar': len([p for p in data['personer'] 
                               if str(p['enhet_id']) == str(e['_id'])])
        } for e in data['enheter']])
        df_enheter.to_excel(writer, sheet_name='Enheter', index=False)
        apply_sheet_styling(writer.sheets['Enheter'], df_enheter)

        # Styrelser & Nämnder
        df_boards = pd.DataFrame([{
            'Namn': b['namn'],
            'Typ': b.get('typ', ''),
            'Förvaltning': b.get('forvaltning_namn', ''),
            'Representanter': ', '.join(b.get('representanter', [])),
            'Ersättare': ', '.join(b.get('ersattare', []))
        } for b in data['boards']])
        df_boards.to_excel(writer, sheet_name='Styrelser & Nämnder', index=False)
        apply_sheet_styling(writer.sheets['Styrelser & Nämnder'], df_boards)

        # Personer
        df_personer = pd.DataFrame([{
            'Namn': p['namn'],
            'Förvaltning': p.get('forvaltning_namn', ''),
            'Avdelning': p.get('avdelning_namn', ''),
            'Enhet': p.get('enhet_namn', ''),
            'Arbetsplats': ', '.join(p.get('arbetsplats', [])),
            'Yrkestitel': p.get('yrkestitel', ''),
            'Telefon': p.get('telefon', ''),
            'Email': p.get('email', ''),
            'Visionombud': 'Ja' if p.get('visionombud') else 'Nej',
            'Skyddsombud': 'Ja' if p.get('skyddsombud') else 'Nej',
            'Huvudskyddsombud': 'Ja' if p.get('huvudskyddsombud') else 'Nej',
            'CSG': f"{p.get('csg_roll', '-') if p.get('csg') else 'Nej'}",
            'LSG/FSG': f"{p.get('lsg_fsg_roll', '-') if p.get('lsg_fsg') else 'Nej'}"
        } for p in data['personer']])
        df_personer.to_excel(writer, sheet_name='Personer', index=False)
        apply_sheet_styling(writer.sheets['Personer'], df_personer)

    # Returnera Excel-filen som bytes
    output.seek(0)
    return output.getvalue()


def show(db):
    """
    Visar nedladdningsknapp för Excel-filen.
    """
    st.header("Exportera Data")
    st.write("Här kan du ladda ner all data från systemet i en Excel-fil.")
    
    # Ladda ner fil
    excel_data = create_excel_file(db)
    st.download_button(
        label="📥 Ladda ner Excel-fil",
        data=excel_data,
        file_name="vision_sektion10_export.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        help="Ladda ner en Excel-fil med all data från systemet"
    )
